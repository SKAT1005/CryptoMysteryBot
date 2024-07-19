import os

import django
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import buttons

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'CryptoMysteryBot.settings')
django.setup()

from const import bot
from app.models import History, User

change_header = [['Дата', 15], ['Вы отдали (валюта)', 20], ['Вы отдали (количество)', 25], ['Вы получили (валюта)', 25],
                 ['Вы получили (количество)', 30], ['Пара обмена', 20], ['Курс валюты, которую отдаете к USDT', 40], ['Курс валюты, которую получаете к USDT', 41]]
buy_header = [['Дата', 15], ['Валюта пополнения', 21], ['Способ пополнения', 25], ['Количество', 20], ['Курс валюты, которую получаете к USDT', 41]]
conclusion_header = [['Дата', 15], ['Валюта вывода', 20], ['Адрес вывода', 25], ['Количество', 30]]
send_to_user_header = [['Дата', 15], ['Валюта перевода', 20], ['Количество', 25],
                       ['Кошелек, на который совершен перевод', 43]]


def headers(sheet, name):
    img = Image("CMB.jpg")
    img.height = 130
    img.width = 128
    sheet.add_image(img, "A1")
    font = Font(size='28', color='333399')
    sheet.cell(row=3, column=3, value=f'История {name} в 👑 Crypto Mystery').font = font


def history(user):
    wb = openpyxl.Workbook()
    change_history = History.objects.filter(user=user, type='Обмен')
    buy_history = History.objects.filter(user=user, type='Пополнения')
    conclusion_history = History.objects.filter(user=user, type='Вывод')
    send_to_user_history = History.objects.filter(user=user, type='Перевод пользователю')
    change = wb.create_sheet(title='Обмены')
    buy = wb.create_sheet(title='Пополнения')
    conclusion = wb.create_sheet(title='Выводы')
    send_to_user = wb.create_sheet(title='Переводы пользователям')
    del wb['Sheet']
    headers(change, 'обменов')
    headers(buy, 'покупупок')
    headers(conclusion, 'выводов')
    headers(send_to_user, 'переводов')
    alf = 'ABCDEFGHI'
    font = Font(size='12', bold=True)
    fill = PatternFill(patternType='solid', fgColor='99ccff')
    bord = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center')
    for col, header in enumerate(change_header, start=1):
        change.column_dimensions[alf[col - 1]].width = header[1]
        cell = change.cell(row=6, column=col, value=header[0])
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = bord

    for col, header in enumerate(buy_header, start=1):
        buy.column_dimensions[alf[col - 1]].width = header[1]
        cell = buy.cell(row=6, column=col, value=header[0])
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = bord

    for col, header in enumerate(conclusion_header, start=1):
        conclusion.column_dimensions[alf[col - 1]].width = header[1]
        cell = conclusion.cell(row=6, column=col, value=header[0])
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = bord

    for col, header in enumerate(send_to_user_header, start=1):
        send_to_user.column_dimensions[alf[col - 1]].width = header[1]
        cell = send_to_user.cell(row=6, column=col, value=header[0])
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = bord

    for row, data in enumerate(change_history, start=7):
        change.cell(row=row, column=1, value=data.date).alignment = alignment
        change.cell(row=row, column=2, value=data.send_cripto).alignment = alignment
        change.cell(row=row, column=3, value=str(data.send_value)).alignment = alignment
        change.cell(row=row, column=4, value=data.get_cripto).alignment = alignment
        change.cell(row=row, column=5, value=str(data.get_value)).alignment = alignment
        change.cell(row=row, column=6, value=f'{data.send_cripto} -> {data.get_cripto}').alignment = alignment
        change.cell(row=row, column=7, value=data.course_send).alignment = alignment
        change.cell(row=row, column=8, value=data.course_get).alignment = alignment

    for row, data in enumerate(buy_history, start=7):
        type = data.get_cripto
        if type == 'RUB':
            type = 'Карта'
        buy.cell(row=row, column=1, value=data.date).alignment = alignment
        buy.cell(row=row, column=2, value=data.get_cripto).alignment = alignment
        buy.cell(row=row, column=3, value=type).alignment = alignment
        buy.cell(row=row, column=4, value=str(data.get_value)).alignment = alignment
        buy.cell(row=row, column=5, value=data.course_get).alignment = alignment

    for row, data in enumerate(conclusion_history, start=7):
        conclusion.cell(row=row, column=1, value=data.date).alignment = alignment
        conclusion.cell(row=row, column=2, value=data.send_cripto).alignment = alignment
        conclusion.cell(row=row, column=3, value=data.address).alignment = alignment
        conclusion.cell(row=row, column=4, value=str(data.send_value)).alignment = alignment

    for row, data in enumerate(send_to_user_history, start=7):
        send_to_user.cell(row=row, column=1, value=data.date).alignment = alignment
        send_to_user.cell(row=row, column=2, value=data.send_cripto).alignment = alignment
        send_to_user.cell(row=row, column=3, value=str(data.send_value)).alignment = alignment
        send_to_user.cell(row=row, column=4, value=data.address).alignment = alignment
    wb.save("История операций.xlsx")
    with open('История операций.xlsx', 'rb') as file:
        bot.send_document(chat_id=user.chat_id, document=file, caption='Файл для скачивания',
                          reply_markup=buttons.go_to_menu())
